from flask import Flask, request, render_template, send_file
import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta
import math, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ---------------- CONFIG ----------------
INDEX_FILE = "IEEMA.xlsx"
INSP_FILE = "insp_call.xlsx"
COEFFS = {"Copper":40,"CRGO":24,"MS":8,"InsMat":4,"TransOil":8,"WPI":8}
TERM_DESC = {
    "Copper":"Copper price index",
    "CRGO":"CRGO steel index",
    "MS":"Mild Steel index",
    "InsMat":"Insulation Material index",
    "TransOil":"Transformer Oil index",
    "WPI":"Wholesale Price Index"
}
CONST_TERM = 8

app = Flask(__name__)

# ---------------- Load Inspection Data ----------------
if not os.path.exists(INSP_FILE):
    raise FileNotFoundError(f"{INSP_FILE} not found.")
insp_df = pd.read_excel(INSP_FILE)
insp_df.columns = [c.strip() for c in insp_df.columns]
if "ICNO" not in insp_df.columns or "call_date" not in insp_df.columns:
    raise ValueError("insp_call.xlsx must have 'ICNO' and 'call_date' columns.")

# ---------------- Helpers ----------------
def parse_date(date_str):
    """Try multiple date formats for input."""
    for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(date_str, fmt)
        except:
            continue
    raise ValueError(f"Invalid date format: {date_str}")

def load_indices():
    if not os.path.exists(INDEX_FILE):
        raise FileNotFoundError(f"{INDEX_FILE} not found.")
    df = pd.read_excel(INDEX_FILE)
    df.columns = [c.strip() for c in df.columns]
    if "Date" not in df.columns:
        possible = [c for c in df.columns if c.lower().startswith("date")]
        if possible:
            df = df.rename(columns={possible[0]: "Date"})
    df["Date"] = pd.to_datetime(df["Date"], dayfirst=True, errors="coerce")
    df = df.dropna(subset=["Date"])
    df = df.set_index(pd.to_datetime(df["Date"].dt.to_period("M").dt.to_timestamp())).sort_index()
    for c in ["C","ES","IS","IM","TO","W"]:
        if c not in df.columns: 
            raise ValueError(f"Column {c} missing in {INDEX_FILE}")
    return df

def get_month_key(dt, prev_month=False):
    if prev_month:
        dt = dt - relativedelta(months=1)
    return datetime(dt.year, dt.month, 1)

def find_index(df, dt, prev_month=False):
    key = get_month_key(dt, prev_month)
    if key not in df.index:
        raise ValueError(f"Indices for {key.strftime('%b-%Y')} not found.")
    return df.loc[key]

def fmt(x, nd=2):
    if isinstance(x,(int,)) or (isinstance(x,float) and x.is_integer()): 
        return f"{int(x):,d}"
    return f"{x:,.{nd}f}"

# ---------------- PVC & LD Calculation ----------------
def calc_pvc(P0, base_date, del_date, df):
    base_row = find_index(df, base_date)
    del_row  = find_index(df, del_date, prev_month=True)
    terms = {}
    mapping = {
        "Copper":("C",base_row["C"],del_row["C"]),
        "CRGO":("ES",base_row["ES"],del_row["ES"]),
        "MS":("IS",base_row["IS"],del_row["IS"]),
        "InsMat":("IM",base_row["IM"],del_row["IM"]),
        "TransOil":("TO",base_row["TO"],del_row["TO"]),
        "WPI":("W",base_row["W"],del_row["W"])
    }
    for t,(col,b,d) in mapping.items():
        b,d = float(b), float(d)
        ratio = d/b if b != 0 else 1
        terms[t] = {"coef":COEFFS[t],"base":b,"del":d,"ratio":ratio,"value":COEFFS[t]*ratio,"desc":TERM_DESC[t]}
    total = CONST_TERM + sum(t["value"] for t in terms.values())
    P = (P0/100)*total
    return {"P0":P0,"base":base_date,"delivery":del_date,"terms":terms,"total":total,"P":P}

def calc_ld(P0, original_dp, delivery_date):
    days = (delivery_date - original_dp).days
    if days <= 0: 
        return {"days":days,"weeks":0,"ld_percent":0,"ld_amt":0}
    weeks = math.ceil(days/7)
    ld_percent = min(weeks/2,10)
    ld_amt = (P0*1.18/100)*ld_percent
    return {"days":days,"weeks":weeks,"ld_percent":ld_percent,"ld_amt":ld_amt}

# ---------------- Excel Report ----------------
def build_excel_report(all_pvcs, final_beneficial, comparisons=None, fname=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison Summary"
    title_font = Font(size=16, bold=True, color="006100")
    hdr_font = Font(bold=True)
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    center = Alignment(horizontal="center")
    ws.merge_cells("A1:E1")
    ws["A1"] = "PVC Beneficial Comparison Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    r = 3
    if comparisons:
        for comp in comparisons:
            ws[f"A{r}"] = comp["title"]; ws[f"A{r}"].font = hdr_font; r+=1
            ws[f"A{r}"] = comp["case1"]["name"]; ws[f"B{r}"] = fmt(comp["case1"]["net"]); r+=1
            ws[f"A{r}"] = comp["case2"]["name"]; ws[f"B{r}"] = fmt(comp["case2"]["net"]); r+=1
            ws[f"A{r}"] = "Beneficial →"; ws[f"B{r}"] = comp["beneficial"]["name"]
            ws[f"B{r}"].font = Font(bold=True, color="006100"); ws[f"B{r}"].fill = green_fill; r+=2
    ws.merge_cells(f"A{r}:E{r}")
    ws[f"A{r}"] = f"FINAL BENEFICIAL CASE: {final_beneficial['name']} | Net PVC: {fmt(final_beneficial['net'])}"
    ws[f"A{r}"].font = Font(size=14, bold=True, color="006100"); ws[f"A{r}"].fill = green_fill; ws[f"A{r}"].alignment = center
    for pvc in all_pvcs:
        ws2 = wb.create_sheet(title=pvc["name"][:25])
        r2 = 1
        ws2[f"A{r2}"] = pvc["name"]; ws2[f"A{r2}"].font = Font(size=14, bold=True); r2 += 2
        ws2[f"A{r2}"] = "Base Date"; ws2[f"B{r2}"] = pvc["base"].strftime("%d-%b-%Y"); r2+=1
        ws2[f"A{r2}"] = "Call/Delivery/Original DP Date"; ws2[f"B{r2}"] = pvc["delivery"].strftime("%d-%b-%Y"); r2+=1
        ws2[f"A{r2}"] = "Quoted Price"; ws2[f"B{r2}"] = fmt(pvc["P0"]); r2+=1
        ws2[f"A{r2}"] = "Term"; ws2[f"B{r2}"] = "Description"; ws2[f"C{r2}"] = "Base"; ws2[f"D{r2}"] = "Delivery"; ws2[f"E{r2}"] = "Ratio"; ws2[f"F{r2}"] = "Coef"; ws2[f"G{r2}"] = "Value"
        for c in ["A","B","C","D","E","F","G"]: ws2[f"{c}{r2}"].font = hdr_font; r2+=1
        for tname,t in pvc["terms"].items():
            ws2[f"A{r2}"]=tname; ws2[f"B{r2}"]=t["desc"]; ws2[f"C{r2}"]=fmt(t["base"]); ws2[f"D{r2}"]=fmt(t["del"])
            ws2[f"E{r2}"]=fmt(t["ratio"],6); ws2[f"F{r2}"]=fmt(t["coef"]); ws2[f"G{r2}"]=fmt(t["value"]); r2+=1
        ws2[f"A{r2}"]="Constant"; ws2[f"G{r2}"]=CONST_TERM; r2+=1
        ws2[f"A{r2}"]="Total Terms Sum"; ws2[f"G{r2}"]=fmt(pvc["total"]); r2+=1
        ws2[f"A{r2}"]="PVC (P)"; ws2[f"G{r2}"]=fmt(pvc["P"]); r2+=1
        if "ld_detail" in pvc:
            r2+=2
            ws2.merge_cells(f"A{r2}:G{r2}")
            ws2[f"A{r2}"]="LIQUIDATED DAMAGES (LD) CALCULATION"
            ws2[f"A{r2}"].font = Font(size=12, bold=True, color="9C0006")
            ws2[f"A{r2}"].fill = red_fill
            ws2[f"A{r2}"].alignment = center; r2+=1
            ld = pvc["ld_detail"]
            ws2[f"A{r2}"]="Delay (days)"; ws2[f"B{r2}"]=ld["days"]; r2+=1
            ws2[f"A{r2}"]="Delay (weeks, ceil)"; ws2[f"B{r2}"]=ld["weeks"]; r2+=1
            ws2[f"A{r2}"]="LD Percent"; ws2[f"B{r2}"]=f"{ld['ld_percent']}%"; r2+=1
            ws2[f"A{r2}"]="LD Amount"; ws2[f"B{r2}"]=fmt(ld["ld_amt"]); r2+=2
            ws2[f"A{r2}"]="Net PVC (after LD deduction)"; ws2[f"B{r2}"]=fmt(pvc.get("net",pvc["P"]))
            ws2[f"A{r2}"].font=Font(bold=True,color="9C0006"); ws2[f"B{r2}"].font=Font(bold=True,color="9C0006")
        else:
            ws2[f"A{r2}"]="Net PVC"; ws2[f"B{r2}"]=fmt(pvc.get("net",pvc["P"]))
        for col in range(1,8):
            ws2.column_dimensions[get_column_letter(col)].width=22
    if not fname:
        fname = f"PVC_Comparison_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(fname)
    return fname

# ---------------- Flask Routes ----------------
@app.route("/", methods=["GET"])
def pvc_form():
    return render_template("pvc_form.html")

@app.route("/calculate", methods=["POST"])
def calculate():
    try:
        P0 = float(request.form["P0"])
        base_date = parse_date(request.form["base_date"])
        original_dp = parse_date(request.form["original_dp"])
        delivery_date = parse_date(request.form["delivery_date"])
        icno = request.form.get("icno","").strip()
        lower = request.form.get("lower","n").lower()

        # --- Call Date Logic ---
        row = insp_df[insp_df["ICNO"].astype(str)==icno]
        if row.empty:
            call_date_str = request.form.get("call_date_manual")
            if not call_date_str:
                return f"<p style='color:red'>Error: ICNO not found and call_date missing</p><a href='/'>Back</a>"
            call_date = parse_date(call_date_str)
        else:
            call_date = pd.to_datetime(row.iloc[0]["call_date"], dayfirst=True, errors='coerce')
            if pd.isna(call_date):
                return f"<p style='color:red'>Call_date invalid in insp_call.xlsx</p><a href='/'>Back</a>"

        df = load_indices()

        # --- PVC & LD Calculation ---
        pvc1 = calc_pvc(P0, base_date, call_date, df); pvc1["name"]="PVC (Actual)"
        ld1 = calc_ld(P0, original_dp, delivery_date)
        pvc1.update({"ld_amt": ld1["ld_amt"], "net": pvc1["P"]-ld1["ld_amt"], "ld_detail": ld1})

        pvc2 = calc_pvc(P0, base_date, original_dp, df); pvc2["name"]="PVC (Contractual)"
        ld2 = calc_ld(P0, original_dp, delivery_date)
        pvc2.update({"ld_amt": ld2["ld_amt"], "net": pvc2["P"]-ld2["ld_amt"], "ld_detail": ld2})

        all_pvcs = [pvc1, pvc2]
        comparisons = []
        benef_normal = pvc1 if pvc1["net"] <= pvc2["net"] else pvc2
        comparisons.append({"title":"Normal PVC Comparison (PVC1 vs PVC2)","case1":pvc1,"case2":pvc2,"beneficial":benef_normal})
        final_beneficial = benef_normal

        # --- Lower Rate ---
        if lower in ("y","yes"):
            P0L = float(request.form["P0_lower"])
            lower_date = parse_date(request.form["lower_date"])
            if original_dp < lower_date:
                pvc_lower = calc_pvc(P0L, lower_date, call_date, df)
                pvc_lower["name"]="Lower Rate PVC (Actual)"; pvc_lower["net"]=pvc_lower["P"]
                all_pvcs.append(pvc_lower)
                final_beneficial = pvc_lower if pvc_lower["net"]<= benef_normal["net"] else benef_normal
            else:
                pvc_a = calc_pvc(P0L, lower_date, original_dp, df); pvc_a["name"]="Lower Rate PVC (Contractual)"; pvc_a["net"]=pvc_a["P"]
                pvc_b = calc_pvc(P0L, lower_date, call_date, df); pvc_b["name"]="Lower Rate PVC (Actual)"; pvc_b["net"]=pvc_b["P"]
                all_pvcs.extend([pvc_a,pvc_b])
                benef_lower = pvc_a if pvc_a["net"]<=pvc_b["net"] else pvc_b
                comparisons.append({"title":"Lower Rate PVC Comparison (Case A vs Case B)","case1":pvc_a,"case2":pvc_b,"beneficial":benef_lower})
                final_beneficial = benef_lower if benef_lower["net"]<= benef_normal["net"] else benef_normal

        fname = build_excel_report(all_pvcs, final_beneficial, comparisons)

        html = "<h3>PVC Calculation Results</h3>"
        for pvc in all_pvcs:
            html += f"<b>{pvc['name']}</b><br>"
            html += f"Base Date: {pvc['base'].strftime('%d-%b-%Y')} | Call Date: {pvc['delivery'].strftime('%d-%b-%Y')}<br>"
            html += f"PVC (P): {fmt(pvc['P'])} | Net PVC: {fmt(pvc.get('net', pvc['P']))}<br><hr>"
        html += f"<h2 style='color:green'>Final Beneficial Case: {final_beneficial['name']} | Net PVC: {fmt(final_beneficial['net'])}</h2><hr>"
        html += f"<a href='/{fname}' download>Download Excel Report</a><br><br><a href='/'>Back</a>"
        return html

    except Exception as e:
        return f"<p style='color:red'>Error: {e}</p><a href='/'>Back</a>"

@app.route("/<filename>")
def download_file(filename):
    return send_file(filename, as_attachment=True)

@app.route("/check_icno")
def check_icno():
    icno = request.args.get("icno","").strip()
    row = insp_df[insp_df["ICNO"].astype(str)==icno]
    if row.empty:
        return {"found": False}
    else:
        call_date = pd.to_datetime(row.iloc[0]["call_date"], dayfirst=True, errors='coerce')
        return {"found": True, "call_date": call_date.strftime("%d-%m-%Y") if not pd.isna(call_date) else ""}

if __name__ == "__main__":
    app.run(debug=True)
